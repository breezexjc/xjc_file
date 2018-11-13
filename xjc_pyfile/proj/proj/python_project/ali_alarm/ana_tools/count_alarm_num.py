from proj.proj.config.database import Postgres
import datetime as dt
import pandas as pd


from openpyxl import load_workbook



import sys, time

class ShowProcess():
    """
    显示处理进度的类
    调用该类相关函数即可实现处理进度的显示
    """
    i = 0 # 当前的处理进度
    max_steps = 0 # 总共需要处理的次数
    max_arrow = 50 #进度条的长度
    infoDone = 'done'

    # 初始化函数，需要知道总共的处理次数
    def __init__(self, max_steps, infoDone = 'Done'):
        self.max_steps = max_steps
        self.i = 0
        self.infoDone = infoDone

    # 显示函数，根据当前的处理进度i显示进度
    # 效果为[>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>]100.00%
    def show_process(self, i=None):
        if i is not None:
            self.i = i
        else:
            self.i += 1
        num_arrow = int(self.i * self.max_arrow / self.max_steps) #计算显示多少个'>'
        num_line = self.max_arrow - num_arrow #计算显示多少个'-'
        percent = self.i * 100.0 / self.max_steps #计算完成进度，格式为xx.xx%
        process_bar = '[' + '>' * num_arrow + '-' * num_line + ']'\
                      + '%.2f' % percent + '%' + '\r' #带输出的字符串，'\r'表示不换行回到最左边
        print(process_bar)
        # 控制台显示
        sys.stdout.write(process_bar) #这两句打印字符到终端
        sys.stdout.flush()
        if self.i >= self.max_steps:
            self.close()

    def close(self):
        print('')
        print(self.infoDone)
        self.i = 0


class AlarmCount():
    pg_inf = {'database': "research", 'user': "postgres", 'password': "postgres",
                   'host': "192.168.20.45", 'port': "5432"}
    def __init__(self):
        self.pg = Postgres.get_instance(AlarmCount.pg_inf)
        self.alarm_count_result = {}
        self.frame = {}

    def data_match(self,month):
        sdate = dt.datetime(year=2018,month=month,day=1)
        # edate = dt.datetime(year=2018,month=month+1,day=1)
        edate = dt.datetime(year=2018, month=month+1, day=1)
        print(sdate,edate)
        sql = "select * from disposal_alarm_data_copy1 where time_point between '{0}' and '{1}' order by time_point"\
            .format(sdate,edate)
        result = self.pg.call_pg_data(sql, fram=True)
        print(result)
        return result

    def alarm_count(self, result):
        result['day'] = result['time_point'].apply(lambda x: x.day)
        result['hour'] = result['time_point'].apply(lambda x: x.hour)
        grouped = result.groupby(['inter_id', 'day'])
        for (k1, k2), group in grouped:
            if k1 not in self.alarm_count_result.keys():
                self.alarm_count_result[k1] ={k2:{}}
            else:
                self.alarm_count_result[k1][k2] = {}
            grouped2 = group.groupby('hour')
            for k3, group2 in grouped2:
                last_alarm = None
                count_alarm = 0
                for i in range(len(group2)):
                    alarm_time = group2.iloc[i][4]
                    if last_alarm:
                        if alarm_time-last_alarm < dt.timedelta(minutes=10):
                            pass
                        else:
                            last_alarm = alarm_time
                            count_alarm += 1
                    else:
                        last_alarm = alarm_time
                        count_alarm += 1
                self.alarm_count_result[k1][k2][k3] = count_alarm

    def frame_create(self):
        alarm_count_result = self.alarm_count_result
        for key1,value1 in alarm_count_result.items():
            int_id = key1
            day = [i + 1 for i in range(30)]
            zero = [0 for i in range(30)]
            index = [i for i in range(24)]
            value_demo = [zero for i in range(24)]
            frame = pd.DataFrame(value_demo, index=index, columns=day)
            self.frame[int_id] = frame

            for key2,value2 in value1.items():

                day = key2

                for key3,value3 in value2.items():
                    hour = key3
                    alarm_times = value3
                    # print(self.frame[int_id])
                    self.frame[int_id].loc[hour][day] = alarm_times
                    # frame.loc[hour]['alarm_times'] = alarm_times
                    # frame = pd.DataFrame(value2, columns=['hour', 'alarm_times'])

    def write_csv(self):
        max_steps = len(self.frame.keys())
        process_bar = ShowProcess(max_steps, 'OK')
        writer = pd.ExcelWriter('alarm_count.xlsx',engine='openpyxl')
        inf = self.call_road_inf()
        try:
            writer.book = load_workbook('alarm_count.xlsx')
            idx = writer.book.sheetnames.index('int_alarm_count')
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet('int_alarm_count', idx)
            writer.save()
        except Exception as e:
            print(e)

        for key, value in self.frame.items():
            process_bar.show_process()
            match_inf = inf[inf['gaode_id'] == key]
            try:
                writer.book = load_workbook('alarm_count.xlsx')
                startrow = writer.book['int_alarm_count'].max_row
                writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
                # writer.book.create_sheet(sheet_name, idx)
            except Exception as e:
                print(e)
                startrow = None

            if not match_inf.empty:
                scats_inf = match_inf.iloc[0][0]
                inter_name = match_inf.iloc[0][2]
                # writer = pd.ExcelWriter('out.xlsx')
                value['Col_sum'] = value.apply(lambda x: x.sum(), axis=1)
                value.loc['Row_sum'] = value.apply(lambda x: x.sum())
                value.insert(0, 'hour', value.index)
                name = [str(scats_inf) + inter_name for i in range(len(value.index))]
                value.insert(0, str(scats_inf) + inter_name, name)
                alarm_sum = value.loc['Row_sum']['Col_sum']
                if alarm_sum > 150:
                    if startrow is None:
                        data = value
                        # idx = writer.book.sheetnames.index('int_alarm_count')
                        # writer.book.create_sheet('int_alarm_count', 0)
                        data.to_excel(writer, 'int_alarm_count', index=False, startrow=0)
                    else:
                        data = value
                        data.to_excel(writer, 'int_alarm_count', index=False, startrow=startrow)
                    writer.save()

        # rexcel = open_workbook('alarm_count.xlsx')  # 用wlrd提供的方法读取一个excel文件
        # rows = rexcel.sheets()[0].nrows  # 用wlrd提供的方法获得现在已有的行数
        # excel = copy(rexcel)  # 用xlutils提供的copy方法将xlrd的对象转化为xlwt的对象
        # table = excel.get_sheet(0)  # 用xlwt对象的方法获得要操作的sheet
        # values = ["1", "2", "3"]
        # row = rows
        # for value in values:
        #     table.write(row, 0, value)  # xlwt对象的写方法，参数分别是行、列、值
        #     table.write(row, 1, "haha")
        #     table.write(row, 2, "lala")
        #     row += 1
        # excel.save("collection.xls")  # xlwt对象的保存方法，这时便覆盖掉了原来的excel

    def call_road_inf(self):
        sql = "select systemid,gaode_id,inter_name from gaode_inter_rel_scats"
        result = self.pg.call_pg_data(sql, fram=True)
        return result

    def fram_merge(self):
        pass

if __name__ == '__main__':

    # x=dt.datetime.strptime('2018-10-01','%Y-%m-%d')
    A = AlarmCount()
    result = A.data_match(10)
    A.alarm_count(result)
    print(A.alarm_count_result)
    A.frame_create()
    print(A.frame)
    A.write_csv()
    # inf = A.call_road_inf()
    # match_inf = inf[inf['gaode_id'] == '14L68097P40']
    # scats_inf = match_inf.iloc[0][0]
    # inter_name = match_inf.iloc[0][2]